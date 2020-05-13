from openpyxl import load_workbook
import argparse, sys, operator
import timeit

start = timeit.default_timer()

parser = argparse.ArgumentParser()
parser.add_argument('data', help='lyric data you wanna compare with specific sentence you will provide')
args = parser.parse_args()

data = args.data
base = input('문장을 입력해주세요!: ')

wb = load_workbook('%s.xlsx'%data)
sentence_sheet = wb.worksheets[0]
max_row = sentence_sheet.max_row

sentence_list = []

for i in range(1, max_row+1):
    sentence_list.append(sentence_sheet['B%d'%i].value)

# for i in sentence_list:
#     print(i)

track_song_info_dict = {}
track_artist_info_dict = {}

# for i in range (1,5):
#     print(len(sentence_list[i]))
    

#n-gram 유사도 비교
def ngram(s, num):
    res = []
    slen = len(s) - num + 1
    for i in range(slen):
        ss = s[i:i+num]
        res.append(ss)
    return res
def diff_ngram(sa, sb, num):
    a = ngram(sa, num)
    b = ngram(sb, num)
    r = []
    cnt = 0
    for i in a:
        for j in b:
            if i == j:
                cnt += 1
                r.append(i)
    return cnt / len(a), r

three_gram_score_list = []
three_gram_word_list = []

for s in sentence_list:
    # 3-gram
    r3, word3  = diff_ngram(base, s, 3)
    three_gram_score_list.append(r3)
    three_gram_word_list.append(word3)


# 3-gram
three_max_index = three_gram_score_list.index(max(three_gram_score_list)) 

tmp_sentence_list = []
tmp_track_list = []

print('\n분석 결과 가장 유사한 문장: %s\n '%sentence_list[three_max_index])

tmp_sentence_list.append(sentence_list[three_max_index])


total = []

try:
    for i in range(0, len(three_gram_score_list)):
        if three_gram_score_list[i] > 0.15:
            if(sentence_list[i] not in tmp_sentence_list ):
                tmp_sentence_list.append(sentence_list[i])
                total.append([three_gram_score_list[i], sentence_list[i]])

    total.sort(key=lambda x:x[0], reverse=True)

    if len(total) >0:
        for t in total:
            print('유사한 문장 순서입니다: %s\n' % (t[1]))

except KeyError:
    sys.exit(1)


stop = timeit.default_timer()
print ('총실행시간은 다음과 같습니다(단위 초): ',stop - start)
